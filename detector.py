import openpyxl

SYSTEM_SHEETS = {'NOMBRE', 'Control', 'MACROS', 'Resumen', 'Resumen_Consolidacion'}

def detect_rem_info(wb, filename=""):
    """
    Identifica la serie de REM (SA, SBM, SBS, SD, SP) y su versión a partir de un workbook openpyxl.
    """
    sheets = set(wb.sheetnames)
    fname_upper = filename.upper() if filename else ""
    
    # 1. Detectar versión desde NOMBRE!A9 si existe
    version = "v1.0"
    if 'NOMBRE' in wb.sheetnames:
        try:
            val_a9 = str(wb['NOMBRE']['A9'].value or '')
            if '1.3' in val_a9:
                version = "v1.3"
            elif '1.2' in val_a9:
                version = "v1.2"
            elif '1.1' in val_a9:
                version = "v1.1"
            elif '1.0' in val_a9:
                version = "v1.0"
        except Exception:
            pass

    # 2. Detectar Serie
    series = "DESCONOCIDO"
    
    # Comprobar por nombres de hojas característicos
    if any(s.startswith('A') and s not in SYSTEM_SHEETS and len(s) <= 4 for s in sheets) or 'SA' in fname_upper:
        series = "SA"
    elif any(s.startswith('BM') for s in sheets) or 'SBM' in fname_upper:
        series = "SBM"
    elif ('B' in sheets or 'B17' in sheets) and not any(s.startswith('BM') for s in sheets):
        series = "SBS"
    elif any(s.startswith('D') and s not in SYSTEM_SHEETS and len(s) <= 4 for s in sheets) or 'SD' in fname_upper:
        series = "SD"
    elif any(s.startswith('P') and s not in SYSTEM_SHEETS and len(s) <= 4 for s in sheets) or 'SP' in fname_upper:
        series = "SP"

    # Hojas de datos destino (excluyendo hojas del sistema)
    target_sheets = [s for s in wb.sheetnames if s not in SYSTEM_SHEETS]
    
    return {
        "series": series,
        "version": version,
        "sheetnames": wb.sheetnames,
        "target_sheets": target_sheets
    }
