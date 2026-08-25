import openpyxl
import io
import os
import unicodedata

SYSTEM_SHEETS = {'NOMBRE', 'Control', 'MACROS', 'Resumen', 'Resumen_Consolidacion'}

def normalize_text(text):
    if not text:
        return ""
    text = str(text).strip().upper()
    text = unicodedata.normalize('NFD', text)
    return ''.join(c for c in text if unicodedata.category(c) != 'Mn')

def is_yellow_input_cell(cell):
    """
    Verifica de forma estricta si una celda corresponde a una celda de ingreso de datos
    (amarillo #FFFFCC / #FFFF99, naranjo suave #FFCC99 / RGB 255,204,153, etc.).
    Retorna False para celdas combinadas (MergedCell), celdas con fórmulas (=) o celdas de encabezados/totales.
    """
    if type(cell).__name__ == 'MergedCell':
        return False
        
    cell_val = cell.value
    if isinstance(cell_val, str) and cell_val.strip().startswith('='):
        return False
        
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
        
    for color_obj in [fill.start_color, fill.fgColor]:
        if not color_obj:
            continue
            
        # 1. Verificación por valor RGB hexadecimal (amarillo #FFFFCC, naranjo #FFCC99 / RGB 255,204,153)
        rgb = getattr(color_obj, 'rgb', None)
        if rgb and isinstance(rgb, str):
            rgb_u = rgb.upper()
            if any(valid_hex in rgb_u for valid_hex in [
                'FFFFCC', 'FFFF99', 'FFFFFFCC', 'FFFFFF99', 
                'FFCC99', 'FFFFCC99', 'FFFFFFCC99',
                'FFF2CC', 'FFFF66', 'FFFF00', 'FFFFFFBA'
            ]):
                return True
                
        # 2. Verificación por Tema de Excel (Tema 4, 5, 6, 7, 9 con tint de ingreso)
        theme = getattr(color_obj, 'theme', None)
        tint = getattr(color_obj, 'tint', None)
        if theme in [4, 5, 6, 7, 9] and tint is not None:
            if 0.3 <= tint <= 0.95 or abs(tint - 0.7999) < 0.2:
                return True
                
        # 3. Verificación por color indexado en paleta estándar (26=light yellow, 43=yellow, 47=orange tint, 13, 34, 65)
        indexed = getattr(color_obj, 'indexed', None)
        if isinstance(indexed, int) and indexed in [26, 43, 47, 13, 34, 65]:
            return True

    return False

def consolidate_rem_files(loaded_items, master_path, series_name):
    """
    Consolida una lista de objetos cargados (con 'name', 'wb_data', 'bytes')
    utilizando el archivo maestro de referencia 'master_path'.
    Modifica ÚNICAMENTE celdas amarillas (#FFFFCC).
    """
    if not master_path or not os.path.exists(master_path):
        raise FileNotFoundError(f"No se encontró la plantilla maestra para la serie {series_name} en {master_path}")

    is_xlsm = master_path.lower().endswith(".xlsm")
    try:
        if is_xlsm:
            out_wb = openpyxl.load_workbook(master_path, keep_vba=True)
        else:
            out_wb = openpyxl.load_workbook(master_path)
    except Exception:
        out_wb = openpyxl.load_workbook(master_path)
        is_xlsm = False

    # Identificar hojas de datos destino
    target_sheets = [s for s in out_wb.sheetnames if s not in SYSTEM_SHEETS]

    for sheet_target in target_sheets:
        out_ws = out_wb[sheet_target]
        start_row = 10 if sheet_target != 'B' else 11
        
        # Mapear Código (Col A o B) o Glosa (Col C o B) a la fila de la plantilla oficial
        code_to_row = {}
        for r in range(start_row, out_ws.max_row + 1):
            c_code = str(out_ws.cell(r, 1).value or out_ws.cell(r, 2).value or '').strip()
            c_glosa = str(out_ws.cell(r, 3).value or out_ws.cell(r, 2).value or '').strip()
            
            if c_code:
                code_to_row[c_code] = r
            elif c_glosa:
                code_to_row[normalize_text(c_glosa)] = r
                
            # RESETEAR A 0 ÚNICAMENTE LAS CELDAS AMARILLAS #FFFFCC
            for c in range(3, out_ws.max_column + 1):
                cell = out_ws.cell(r, c)
                if is_yellow_input_cell(cell):
                    cell.value = 0

        # ACUMULAR DATOS DE CADA ARCHIVO DE LA SERIE
        for item in loaded_items:
            src_wb = item["wb_data"]
            if sheet_target not in src_wb.sheetnames:
                continue
            src_ws = src_wb[sheet_target]
            src_start = 10 if sheet_target != 'B' else 11
            
            for r in range(src_start, src_ws.max_row + 1):
                c_code = str(src_ws.cell(r, 1).value or src_ws.cell(r, 2).value or '').strip()
                c_glosa = str(src_ws.cell(r, 3).value or src_ws.cell(r, 2).value or '').strip()
                
                target_row = None
                if c_code and c_code in code_to_row:
                    target_row = code_to_row[c_code]
                elif c_glosa and normalize_text(c_glosa) in code_to_row:
                    target_row = code_to_row[normalize_text(c_glosa)]
                    
                if target_row:
                    for c in range(3, src_ws.max_column + 1):
                        target_cell = out_ws.cell(target_row, c)
                        if is_yellow_input_cell(target_cell):
                            val = src_ws.cell(r, c).value
                            if isinstance(val, (int, float)) and val != 0:
                                curr = target_cell.value
                                curr_num = curr if isinstance(curr, (int, float)) else 0
                                target_cell.value = curr_num + val

    # Preservar datos de encabezado desde el primer archivo cargado si existe hoja NOMBRE
    if 'NOMBRE' in out_wb.sheetnames and loaded_items:
        try:
            first_src_wb = loaded_items[0]["wb_data"]
            if 'NOMBRE' in first_src_wb.sheetnames:
                src_name_ws = first_src_wb['NOMBRE']
                out_name_ws = out_wb['NOMBRE']
                # Copiar datos principales de celda (Establecimiento, Comuna, Servicio)
                for cell_ref in ['B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7']:
                    val = src_name_ws[cell_ref].value
                    if val:
                        out_name_ws[cell_ref].value = val
        except Exception:
            pass

    # Eliminar hoja de resumen previa si existe
    if "Resumen_Consolidacion" in out_wb.sheetnames:
        out_wb.remove(out_wb["Resumen_Consolidacion"])

    output_stream = io.BytesIO()
    out_wb.save(output_stream)
    output_stream.seek(0)
    
    out_filename = f"Consolidado_REM_{series_name}_2026.xlsm" if is_xlsm else f"Consolidado_REM_{series_name}_2026.xlsx"
    out_mime = "application/vnd.ms-excel.sheet.macroEnabled.12" if is_xlsm else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return {
        "bytes": output_stream.getvalue(),
        "filename": out_filename,
        "mime": out_mime,
        "series": series_name,
        "is_xlsm": is_xlsm
    }
